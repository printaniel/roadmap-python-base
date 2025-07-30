import logging.config
import openpyxl
import os
from rich import print
from rich.console import Console
from rich .table import Table
from openpyxl import Workbook, load_workbook
import msvcrt
from Dia_7 import excepciones
import logging 

logging.basicConfig(
    level = logging.INFO,
    format = '{asctime} | {levelname}: {message}',
    style = '{',
    datefmt= '%Y-%m-%d %H:%M:%S',
    handlers= [logging.FileHandler(filename = "app_log", mode = 'a', encoding= "utf-8")]

)


archivo = "inventario_producto.xlsx"
salida = False

def main():
    
    logging.info("Inicio de la aplicación")
    while not salida:
        mostrar_tabla()
        eleccion_de_opcion()
        borrar_datos_consola()
        
    return 0

# creasion e impresion de una tabla de Gestion de Productos 
def mostrar_tabla():
    
    consola = Console()
    table = Table(title = "Gestión de Productos")
    
    table.add_column("Acciones", justify="center", header_style="bold blue")
    table.add_row("1. Agregar producto")
    table.add_row("2. Eliminar producto")
    table.add_row("3. Actualizar producto")
    table.add_row("4. Mostrar todos los productos")
    table.add_row("5. Buscar producto por su nombre")
    table.add_row("6. Salir")
    
    consola.print(table)

# Pedir una opcion al usuario
def pedir_opcion():
    opcion = input("Elija una opcion: ")
    return opcion

# eleccion de una opcion por parte del usuario
def eleccion_de_opcion():
    
    opcion = ""
    opciones = {"1": registrar_producto,"2": eliminar_producto ,"3": actualizar_producto,"4": mostrar_todos_productos,
                "5": buscar_producto, "6": salir}

    print(opcion)
    while not opcion in opciones.keys():
        opcion = pedir_opcion()
        if opcion in opciones.keys():
            opciones[opcion]()
        else:
            print("[red]ERROR: [/red]Opcion incorrecta")
            
def salir():
    global salida
    salida = True
    print("[red]Has salido del programa[/red]")
    logging.info("Termina el programa")

def registrar_producto():
    global archivo
    nombre,precio,cantidad_stock = pedir_datos_producto()
    id = 0
    if os.path.exists(archivo):
        inventario = load_workbook(archivo)
        hoja = inventario.active
        id = hoja.max_row
        
    else:
        inventario = Workbook()
        hoja = inventario.active
        hoja.title = "Inventario"
        hoja.append(["ID","Nombre","Precio", "Stock"])
        id = hoja.max_row
        
    hoja.append([id,nombre,precio,cantidad_stock])      
    try:
        inventario.save(archivo)
        producto_registrado_log()
    except PermissionError as e:
        inventario_en_uso_log()
        raise excepciones.PermisoEdicionDenegado() from e

    

def eliminar_producto():
    global archivo
    
    try:   
        if os.path.exists(archivo):
            
            inventario = load_workbook(archivo)
            hoja = inventario.active
            
            if hoja.max_row > 1:
                nombre = pedir_nombre_con_duplicados().strip()
                fila = buscar_fila_producto(nombre)            

                if fila != None:
                    producto_encontrado_log()
                    hoja.delete_rows(fila)                   
                    inventario.save(archivo)
                    producto_eliminado_log()
                    print("[green]Éxito: [/green] El producto ha sido eliminado correctamente")
                                       
                else:
                    producto_no_encontrado_log()
                    raise excepciones.ProductoNoEncontrado(nombre)
            else:
                inventario_vacio_log()
                raise excepciones.InventarioVacioError()
            
        else:
            inventario_inexistente_log()
            raise excepciones.InventarioNoEncontradoError()
    except PermissionError as e:
        inventario_en_uso_log()
        raise excepciones.PermisoEdicionDenegado() from e
    
    except Exception as e :
        print(f"[red]ERROR: [/red]{e}")
    
def buscar_fila_producto(nombre):
    global archivo
    
    fila = None
    if os.path.exists(archivo):
        registro = load_workbook(archivo)
        hoja = registro.active
        
        for producto in hoja.iter_rows(min_row = 2,values_only = False):
            if producto[1].value == nombre.strip():
                fila = producto[1].row
        
    return fila
    

def pedir_datos_producto():
    datos_producto = []
    
    datos_producto.append(pedir_nombre())
    datos_producto.append(pedir_precio())
    datos_producto.append(pedir_cantidad_stock())
    
    return datos_producto


# Peticiones al usuario
def pedir_nombre():
    es_valido = False
    
    
    while not es_valido:
        try:
            nombre = input("Escriba el nombre: ")
            validar_nombre(nombre)
            es_valido = True 
        except Exception as e:
            datos_invalidos_log()
            print(f"[red]ERROR: [/red] {e}")
            
    return nombre
def pedir_nombre_con_duplicados():
    es_valido = False
      
    while not es_valido:
        try:
            nombre = input("Escriba el nombre: ")
            validar_nombre_con_duplicados(nombre)
            es_valido = True 
        except Exception as e:
            datos_invalidos_log()
            print(f"[red]ERROR: [/red] {e}")
            
    return nombre
    
def pedir_precio():
    es_valido = False
    
    while not es_valido:
        try:
            precio = input("Escriba el precio: ")
            validar_precio(precio)
            es_valido = True 
        except Exception as e:
            datos_invalidos_log()
            print(f"[red]ERROR: [/red] {e}")
            
    return precio
    
def pedir_cantidad_stock():
    
    es_valida = False
    
    while not es_valida:
        try:
            cantidad_stock = input("Escriba la cantidad en stock: ")
            validar_cantidad_stock(cantidad_stock)
            es_valida = True 
        except Exception as e:
            datos_invalidos_log()
            print(f"[red]ERROR: [/red] {e}")
            
    return cantidad_stock
    
    
# Validaciones 
def validar_nombre(nombre):
    global archivo
    
    nombre = nombre.strip()
    
    if len(nombre) == 0:
        
        raise excepciones.NombreVacio()
     
    for letra in nombre:
        if not letra.isalpha() and not letra == " ":
            raise excepciones.NombreAlfabeticoInvalido(nombre)

        
    # Comprobar duplicados
    if os.path.exists(archivo):
        
        inventario = load_workbook(archivo)
        hoja = inventario.active
        
        fila = buscar_fila_producto(nombre)
        if fila != None:
            raise excepciones.NombreDuplicado(nombre)
           
    return nombre   

def validar_nombre_con_duplicados(nombre):
    global archivo
    
    nombre = nombre.strip()
    
    if len(nombre) == 0:
        raise ValueError("El nombre no puede estar vacio")
     
    for letra in nombre:
        if not letra.isalpha() and not letra == " ":
            raise excepciones.NombreAlfabeticoInvalido(nombre)         
    
def validar_precio(precio):
    precio = precio.strip()
    es_numero = False
    
    try: 
        precio = float(precio)
        es_numero = True
    except:
        raise excepciones.ValorNoNumerico()
    
    if es_numero:
        if precio < 0:
            raise excepciones.NumeroNoPositivo()
        
    return precio
    
    
def validar_cantidad_stock(numero):
    numero = numero.strip()
    es_numero = False
    try:
        numero = int(numero)
        es_numero = True
    except:
        raise excepciones.NumeroNoEntero()
    
    if es_numero:
        if numero < 0:
            raise excepciones.NumeroNoPositivo()

 
def mostrar_todos_productos():
    global archivo
    try: 
        if os.path.exists(archivo):       
            inventario = load_workbook(archivo)
            hoja = inventario.active
            contador = 1
            
            if hoja.max_row > 1:
                print("\n")
                for producto in hoja.iter_rows(values_only = False, min_row = 2):
                    print(f"Producto #{contador}")
                    print(f"ID: {producto[0].value}")
                    print(f"Nombre: {producto[1].value}")
                    print(f"Precio: {producto[2].value}")
                    print(f"Cantidad en stock: {producto[3].value}")
                    print("\n")
                    contador += 1
            else:
                inventario_vacio_log()
                raise excepciones.InventarioVacioError()
                
        else:
            inventario_inexistente_log()
            raise excepciones.InventarioNoEncontradoError()  
    except Exception as e:
        print(f"[red]ERROR: [/red]{e}") 

def buscar_producto():
    
    try:
        if os.path.exists(archivo):       
            inventario = load_workbook(archivo)
            hoja = inventario.active   
            encontrado = False     
            if hoja.max_row > 1:
                
                nombre = pedir_nombre_con_duplicados().strip()
                
                for producto in hoja.iter_rows(values_only = False, min_row = 2): 
                    if nombre == producto[1].value:
                        print("\n")
                        print(f"[greeen]Producto Encontrado:[/green] ")
                        print(f"ID: {producto[0].value}")
                        print(f"Nombre: {producto[1].value}")
                        print(f"Precio: {producto[2].value}")
                        print(f"Cantidad en stock: {producto[3].value}")
                        print("\n")
                        encontrado = True
                if not encontrado:
                    producto_no_encontrado_log()
                    raise excepciones.ProductoNoEncontrado(nombre)
                    
            else:
                inventario_vacio_log()
                raise excepciones.InventarioVacioError()
            
        else:
            inventario_inexistente_log()
            raise excepciones.InventarioNoEncontradoError()
    except PermissionError as e:
        inventario_en_uso_log()
        raise excepciones.PermisoEdicionDenegado() from e

    except Exception as e: 
        e = str(e).replace("[", "[[").replace("]", "]]")  # Escapa corchetes
        print(f"[red]ERROR: [/red] {e}")
    

def actualizar_producto():
    
    try:
        if os.path.exists(archivo):       
            inventario = load_workbook(archivo)
            hoja = inventario.active   
            actualizado = False
            
            if hoja.max_row > 1:
                
                print("[blue]INFO:[/blue]Se necesita el nombre del producto a buscar: ")
                nombre = pedir_nombre_con_duplicados().strip()
                
                for producto in hoja.iter_rows(values_only = False, min_row = 2): 
                    if nombre == producto[1].value:
                        print("[blue]INFO: [/blue]Escriba los nuevos datos para el producto a actualizar")          
                        producto[1].value ,producto[2].value, producto[3].value = pedir_datos_producto() 
                        actualizado = True
                        producto_actualizado_log()
                        print(f"[green]ÉXITO:[/green]Se ha actualizado el producto exitosamente")       
                
                if not actualizado:
                    producto_no_encontrado_log()
                    raise excepciones.ProductoNoEncontrado(nombre)
                    
                inventario.save(archivo)               
            else:
                inventario_vacio_log()
                raise excepciones.InventarioVacioError()
            
        else:
            inventario_inexistente_log()
            raise excepciones.InventarioNoEncontradoError()
    except PermissionError as e:
        inventario_en_uso_log()
        raise excepciones.PermisoEdicionDenegado() from e
    
    except Exception as e:
        print(f"[red]ERROR: [/red]{e}")
        
def borrar_datos_consola():
    global salida
    
    if not salida:    
        print("Presiona cualquier tecla para continuar...")
        tecla = msvcrt.getch()
        if tecla != None:
            os.system('cls' if os.name == 'nt' else 'clear')

def datos_invalidos_log():
    logging.error("El usuario ha enviado datos invalidos ")
    
def inventario_inexistente_log():
    logging.error("El usuario intenta usar el registro no creado")
    
def inventario_vacio_log():
    logging.error("El usuario intenta hacer una accion en un inventario vacio")

def inventario_en_uso_log():
    logging.error("El usuario no tiene permisos para editar el inventario")
    
def producto_no_encontrado_log():
    logging.error("El usuario intentó buscar un producto que no existe")

def producto_registrado_log():
    logging.info("Se ha registrado un nuevo producto")

def producto_encontrado_log():
    logging.info("Se ha encontrado el producto buscado")

def producto_eliminado_log():
    logging.info("Se ha eliminado un producto correctamente ")

def producto_actualizado_log():
    logging.info("Producto actualizado")


main()
