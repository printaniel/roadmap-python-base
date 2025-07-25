"""✅ 2. Ejercicio integrador: sistema simple de notas
Enunciado:
Crea un programa que permita:
Registrar estudiantes (nombre, ID, materia)
Guardar sus notas (puedes usar diccionarios con listas)
Calcular promedio por estudiante
Guardar y leer los datos desde un archivo

Requisitos técnicos:
Usa funciones para cada operación
Usa estructuras como diccionarios, listas y archivos
Interfaz por consola tipo menú (1. Agregar, 2. Calcular promedio, 3. Guardar, 4. Salir)"""

from rich import print
from rich.table import Table
from rich.console import Console
from datetime import time
from openpyxl import Workbook, load_workbook
import os
import msvcrt

from rich.table import Table
from rich.console import Console


lista_estudiantes=[]
archivo = "Registro Estudiantes.xlsx"
salida = False


def mostrar_tabla():
    consola = Console()

    # Crear tabla con título en negrita y color
    table = Table(
        title="[bold blue]📋 Menú de Gestión de Estudiantes[/bold blue]",
        show_lines=True,  # Añade líneas horizontales
        header_style="bold white",
        border_style="bright_blue"
    )

    # Agregar columna con alineación centrada
    table.add_column("Acciones", justify="center", header_style="bold blue")

    # Agregar filas con estilo variado
    table.add_row("1. 📄 [bold]Mostrar Lista de Estudiantes[/bold]")
    table.add_row("2. 📝 [blue]Registrar nuevo Estudiante[/blue]")
    table.add_row("3. ❌ [bold]Eliminar un estudiante[/bold]")
    table.add_row("4. 📊 [bold]Calcular Promedio de un Estudiante[/bold]")
    table.add_row("5. 📈 [bold]Calcular Promedio Total[/bold]")
    table.add_row("6. 🚪 [bold red]Salir[/bold red]")

    # Mostrar la tabla
    consola.print(table)

def registrar_estudiante():
    global archivo
    try:
        # Pedir datos
        nombre = pedir_nombre()  
        id = pedir_id()
        materia = pedir_materia()
        notas = pedir_notas_estudiante()
        # Registro
        
        if os.path.exists(archivo):
            # Se carga el registro en caso de que no exista
           libro = load_workbook(archivo)
           hoja = libro.active
        else:
            # Se crea uno nuevo en caso de que no exista
            libro = Workbook()
            hoja = libro.active
            hoja.title = "Estudiantes"
            hoja.append(["Nombre","ID", "Materia", "Nota #1", "Nota #2", "Nota #3", "Nota #4", "Nota #5"])
        # Agregar los datos del nuevo estudiant
        hoja.append([nombre, id, materia , *notas])   
        # Guardar los cambios
        libro.save(archivo)  
    except ValueError: 
        print("[red]ERROR:[/red] El estudiante no ha sido registrado (problemas en los datos de inscripcion)")
    else:
        print("[blue]INFO:[/blue] Se ha guardado el estudiante correctamente")
 
 ## Peticion de datos 
 
 #Pedir notas de estudiante
def pedir_notas_estudiante():
    # Asumiendo que un estudiante tenga  5 notas por materia
    cantidad_notas = 5
    notas = []
    
    print(f"A continuacion se pediran las {cantidad_notas} notas del estudiante...")
    while len(notas) < 5:
        try:
            nota = int(input(f"Digite la nota #{len(notas)+1} del estudiante: "))
            validar_nota(nota)
            notas.append(nota)
            
        except ValueError:
            print(f"[red]ERROR[/red]: Las notas deben ser un valor numerico")
        except Exception as e:
            print(f"[red]ERROR[/red]: {e}")
           
    return notas    
    
# Pedir nombre del estudiante  
def pedir_nombre():
    
    es_valido = False
    while not es_valido:
        try:
            nombre = input("Escriba el nombre del estudiante: ")
            validar_nombre(nombre)
            es_valido = True 
            
        except ValueError as e:
            print(f"[red]ERROR[/red]: {e}")
    return nombre

# Pedir Materia del estudiante
def pedir_materia():
    
    es_valido = False
    while not es_valido:
        try:
            materia = input("Escriba el nombre de la materia del estudiante [Matematica, Historia, Español]: ")
            validar_materia(materia)
            es_valido = True 
            
        except ValueError as e:
            print(f"[red]ERROR[/red]: {e}")
    
    return materia

# Pedir ID del estudiante
def pedir_id():
    
    es_valido = False
    while not es_valido:
        try:
            id = input("Escriba el ID del estudiante: ")
            id = id.strip()
            validar_id(id)
            
            es_valido = True 
            
        except ValueError as e:
            print(f"[red]ERROR[/red]: {e}")
    return id

# Pedir ID del estudiante
def pedir_id_sin_validar_duplicados():
    
    es_valido = False
    while not es_valido:
        try:
            id = input("Escriba el ID del estudiante: ")
            id = id.strip()
            
            validar_id_con_duplicados(id)
            es_valido = True 
            
        except ValueError as e:
            print(f"[red]ERROR[/red]: {e}")
    return id
     
   
    
## Validacion de nombre, id, materia y notas
def validar_nombre(nombre):
    es_valido = True
    
    if len(nombre.strip()) == 0:
        raise ValueError("Nombre incorrecto")
    for i in nombre:
        if not i.isalpha() and not i ==" ":
            es_valido = False
            raise ValueError("Nombre incorrecto")
    return es_valido



# Arreglar validacion de carnet a String
def validar_id(id):
    es_valido = True
    global lista_estudiantes
    global archivo
    
    
    if os.path.exists(archivo):
         # Se carga el registro en caso de que no exista
        libro = load_workbook(archivo)
        hoja = libro.active
    else:
            # Se crea uno nuevo en caso de que no exista
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Estudiantes"
        hoja.append(["Nombre","ID", "Materia", "Nota #1", "Nota #2", "Nota #3", "Nota #4", "Nota #5"])
    
    
    for estudiante in hoja.iter_rows(min_row=2, values_only= False):
        if estudiante[1].value == id:
            raise ValueError("Ya existe un estudiante con ese id")
    
    if not len(id) == 11 or not validar_string_de_numeros(id):
        es_valido = False
        raise ValueError("El id debe contener 11 numeros ")
    return es_valido

def validar_id_con_duplicados(id):
    es_valido = True
    global lista_estudiantes

    if not len(id) == 11 or not validar_string_de_numeros(id):
        es_valido = False
        raise ValueError("El id debe contener 11 numeros ")
    return es_valido





def validar_materia(materia: str):
    es_valido = True
    
    lista_materias = ["Matematicas","Historia","Español"]
    materia = materia.lower().capitalize()
    if not materia.strip() in lista_materias:
        es_valido = False
        raise ValueError("Materia incorrecta")
    return es_valido

def validar_nota(nota: int):
    es_valido = True
    if not isinstance(nota, int) or nota > 5 or nota < 2 :
        es_valido = False
        raise Exception("La nota debe ser un numero entre 2 y 5")
    return es_valido

def validar_string_de_numeros(cadena_numeros):
    es_valido = True
    try: 
        for i in cadena_numeros:
            numero = int(i)
    except:
        es_valido = False
        
    return es_valido
        
   
def eliminar_estudiante():
    global archivo
            
    if os.path.exists(archivo):
        
        registro = load_workbook(archivo)
        hoja = registro.active
        
        if hoja.max_row < 2:
            print("[blue]INFO: [/blue]No hay estudiantes en el registro")
        else:
            id = pedir_id_sin_validar_duplicados()
                       
            fila_eliminar = None
            
            for estudiante in hoja.iter_rows(min_row = 2, values_only = False):
                if estudiante[1].value.strip() == id.strip():
                    fila_eliminar = estudiante[1].row
                    break # se sale de la busqueda al encontrar el estudiante
            
            if fila_eliminar != None:
                print(fila_eliminar)
                hoja.delete_rows(fila_eliminar)
                registro.save(archivo)
                print("[blue]INFO: [/blue]Se ha eliminado al estudiante correctamente")
            else:
                print("[red]ERROR: [/red]No se ha encontrado el estudiante")
            
    else:
            print("[red]ERROR: [/red]No existe el Registro de Estudiantes ")
            

def calcular_promedio_Total():
    
    global archivo
    total = 0
    
    if os.path.exists(archivo):
        
        registro = load_workbook(archivo)
        hoja = registro.active
        
        if hoja.max_row < 2:
            print("[blue]INFO: [/blue]No hay estudiantes en el registro")
        else:                       
            for estudiante in hoja.iter_rows(min_row = 2,values_only = False):
                total += estudiante[3].value + estudiante[4].value + estudiante[5].value + estudiante[6].value + estudiante[7].value   
                
            return (total / 5) /(hoja.max_row - 1)
    else:
        print("[red]ERROR: [/red]No existe el Registro de Estudiantes ")
    

   

def calcular_promedio_estudiante():
        
    encontrado = False
    total = 0
    
    if os.path.exists(archivo):
        registro = load_workbook(archivo)
        hoja = registro.active
        
        if hoja.max_row < 2:
            print("[blue]INFO: [/blue]No hay estudiantes en el registro")
        else:
            id = pedir_id_sin_validar_duplicados()
            for estudiante in hoja.iter_rows(min_row = 2,values_only = False):
                if estudiante[1].value == id:
                    encontrado = True
                    total += estudiante[3].value + estudiante[4].value + estudiante[5].value + estudiante[6].value + estudiante[7].value
        
            if not encontrado:
                print("[red]ERROR: [/red]No se ha encontrado el estudiante") 
            
    else:
        print("[red]ERROR: [/red]No existe el Registro de Estudiantes ")
    
    return total / 5
    

def mostrar_lista_estudiantes():
    global archivo
    
    contador = 0
    if os.path.exists(archivo):
        registro = load_workbook(archivo)
        hoja = registro.active
        
        if hoja.max_row < 2:
            print("[blue]INFO: [/blue]No hay estudiantes en el registro")
        else:
          
                for estudiante in hoja.iter_rows(min_row = 2,values_only = False):
                    print(f"Estudiante #{contador + 1}")   
                    print(f"Nombre: {estudiante[0].value}")
                    print(f"ID: {estudiante[1].value}")
                    print(f"Materia: {estudiante[2].value}")
                    print(f"Notas: [{estudiante[3].value, estudiante[4].value , estudiante[5].value ,
                        estudiante[6].value , estudiante[7].value }]")          
                    print("\n")
                    contador += 1
            
    else:
        print("[red]ERROR: [/red]No existe el Registro de Estudiantes ")
    
def pedir_opcion():
    
    opcion = input("Elija una opcion: ")
    
    match opcion:
        case "1":
            mostrar_lista_estudiantes()
        case "2":
            registrar_estudiante()
        case "3":
            eliminar_estudiante()          
        case "4":
            promedio = calcular_promedio_estudiante()
            if promedio != None and promedio != 0:
                print(f"Promedio del estudiante: {promedio}")
            
        case "5":
            promedio_total = calcular_promedio_Total()
            if  not promedio_total == None and not promedio_total == 0:
                print(f"Promedio total de los Estudiantes {promedio_total}")
        case "6":
            salir()
        case _:
            print("[red]ERROR: [/red]Opción no válida")
   
   


 
def salir():
    global salida
    salida = True
    print("Has salido del programa")
 
def borrar_datos_consola():
    global salida
    
    if not salida:    
        print("Presiona cualquier tecla para continuar...")
        tecla = msvcrt.getch()
        if tecla != None:
            os.system('cls' if os.name == 'nt' else 'clear')

    
       
def main():
    global salida
    
    while salida == False:
        mostrar_tabla()
        pedir_opcion()
        borrar_datos_consola()
        
    

main()

    